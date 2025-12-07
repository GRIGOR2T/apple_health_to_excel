from pathlib import Path
from lxml import etree as ET
import datetime as dt
import pandas as pd
import numpy as np
from tqdm import tqdm
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

XML_FILE = Path("export.xml")
START_DATE = dt.date(2025, 8, 10)


def output_filename():
    now = dt.datetime.now().strftime("%Y%m%d_%H%M")
    return f"weight_vo2_history_{now}.xlsx"


def parse_dt(s: str) -> dt.datetime:
    # "2025-11-30 12:51:18 +0100" -> datetime
    return dt.datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")


def parse_export(xml_path: Path):
    weight_rows = []
    vo2_rows = []

    print("Читаю XML:", xml_path.name)
    context = ET.iterparse(str(xml_path), events=("end",), tag="Record")

    for _, elem in tqdm(context, desc="Парсинг", unit="rec"):
        r_type = elem.get("type")
        value = elem.get("value")
        start = elem.get("startDate")

        try:
            val_float = float(value)
        except Exception:
            elem.clear()
            continue

        dt_start = parse_dt(start)
        if dt_start.date() < START_DATE:
            elem.clear()
            continue

        if r_type == "HKQuantityTypeIdentifierBodyMass":
            weight_rows.append((dt_start.date(), val_float))
        elif r_type == "HKQuantityTypeIdentifierVO2Max":
            vo2_rows.append((dt_start.date(), val_float))

        elem.clear()

    del context
    return (
        pd.DataFrame(weight_rows, columns=["date", "weight_kg"]),
        pd.DataFrame(vo2_rows, columns=["date", "vo2max"]),
    )


def aggregate_daily(df_weight, df_vo2):
    df_w = df_weight.groupby("date", as_index=False)["weight_kg"].mean()
    df_v = df_vo2.groupby("date", as_index=False)["vo2max"].max()

    daily = (
        pd.merge(df_w, df_v, on="date", how="outer")
        .sort_values("date")
        .reset_index(drop=True)
    )
    return daily


def add_metrics(daily: pd.DataFrame):
    daily["date"] = pd.to_datetime(daily["date"], errors="coerce")

    daily["days_from_start"] = (daily["date"] - daily["date"].iloc[0]).dt.days

    start_weight = daily["weight_kg"].dropna().iloc[0]
    start_vo2 = daily["vo2max"].dropna().iloc[0]

    daily["weight_delta_kg_from_start"] = daily["weight_kg"] - start_weight
    daily["vo2_delta_from_start"] = daily["vo2max"] - start_vo2

    daily["vo2_gain_per_kg_lost"] = np.where(
        (daily["weight_delta_kg_from_start"] < 0)
        & daily["vo2_delta_from_start"].notna(),
        daily["vo2_delta_from_start"] / (-daily["weight_delta_kg_from_start"]),
        np.nan,
    )

    return daily


def save_excel(daily: pd.DataFrame):
    cols = [
        "date",
        "weight_kg",
        "vo2max",
        "days_from_start",
        "weight_delta_kg_from_start",
        "vo2_delta_from_start",
        "vo2_gain_per_kg_lost",
    ]

    final = (
        daily[cols]
        .sort_values("date", ascending=False)
        .reset_index(drop=True)
    )

    float_cols = [
        "weight_kg",
        "vo2max",
        "weight_delta_kg_from_start",
        "vo2_delta_from_start",
        "vo2_gain_per_kg_lost",
    ]
    final[float_cols] = final[float_cols].round(2)

    final = final.rename(
        columns={
            "date": "Дата",
            "weight_kg": "Вес, кг",
            "vo2max": "VO₂max",
            "days_from_start": "Дней от старта",
            "weight_delta_kg_from_start": "Δ веса от старта, кг",
            "vo2_delta_from_start": "Δ VO₂max от старта",
            "vo2_gain_per_kg_lost": "Прирост VO₂max на 1 кг снижения веса",
        }
    )

    out_name = output_filename()
    print("Сохраняю в Excel:", out_name)

    with pd.ExcelWriter(out_name, engine="openpyxl") as writer:
        final.to_excel(writer, sheet_name="Weight_vs_VO2", index=False)
        ws = writer.sheets["Weight_vs_VO2"]

        ws.freeze_panes = "A2"

        for cell in ws["A"][1:]:
            if cell.value:
                cell.number_format = "mm/dd/yyyy"

        for col_idx, column in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in column:
                cell.font = Font(size=22)

                val = cell.value
                if val is None:
                    continue
                if isinstance(val, dt.datetime):
                    text = val.strftime("%m/%d/%Y")
                else:
                    text = str(val)
                if len(text) > max_len:
                    max_len = len(text)

            width = max_len * 1.4 + 6
            ws.column_dimensions[get_column_letter(col_idx)].width = width

def main():
    print("Стартовая дата:", START_DATE)
    df_weight, df_vo2 = parse_export(XML_FILE)
    daily = aggregate_daily(df_weight, df_vo2)
    daily = add_metrics(daily)
    print("Строк после агрегации:", len(daily))
    save_excel(daily)
    print("Готово.")


if __name__ == "__main__":
    main()