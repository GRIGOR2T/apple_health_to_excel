from pathlib import Path
from lxml import etree as ET
import datetime as dt
import pandas as pd
import sys
import time

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

EXPORT_XML = Path("export.xml")
CUTOFF_DATE = dt.date(2025, 8, 10)  # только с 10.08.2025 и новее


def parse_dt(s: str) -> dt.datetime:
    # Формат Apple: "2025-11-28 06:18:00 +0100"
    return dt.datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")


def format_vo2(value: float) -> str:
    # До 4 знаков после запятой, без лишних нулей
    s = f"{value:.4f}"
    s = s.rstrip("0").rstrip(".")
    return s


def extract_vo2_since_cutoff():
    """Стримим export.xml и вытаскиваем VO2max записи >= CUTOFF_DATE."""
    rows = []

    if not EXPORT_XML.exists():
        raise FileNotFoundError(f"{EXPORT_XML} не найден")

    print(f"Читаю {EXPORT_XML} …")
    print(f"Берём записи начиная с: {CUTOFF_DATE}")

    context = ET.iterparse(str(EXPORT_XML), events=("end",), tag="Record")

    processed = 0
    found = 0
    spinner = ["|", "/", "-", "\\"]
    start_time = time.time()

    for _, elem in context:
        processed += 1

        # Обновляем статус каждые 100k элементов
        if processed % 100000 == 0:
            elapsed = time.time() - start_time
            speed = processed / elapsed if elapsed > 0 else 0
            spin = spinner[(processed // 100000) % 4]
            sys.stdout.write(
                f"\r{spin} Обработано: {processed:,} | VO2max найдено: {found:,} | ~{speed:,.0f}/сек"
            )
            sys.stdout.flush()

        if elem.get("type") == "HKQuantityTypeIdentifierVO2Max":
            found += 1

            val_str = elem.get("value")
            vo2 = float(val_str)  # уже ml/kg/min

            dt_start = parse_dt(elem.get("startDate"))
            d = dt_start.date()

            if d >= CUTOFF_DATE:
                rows.append((d, vo2))

        elem.clear()

    del context

    elapsed = time.time() - start_time
    sys.stdout.write(
        f"\r✓ Готово. Обработано: {processed:,} | VO2max найдено: {found:,} | Время: {elapsed:.1f} c\n"
    )
    sys.stdout.flush()

    return rows


def main():
    rows = extract_vo2_since_cutoff()

    if not rows:
        print("Не найдено VO2max после 2025-08-10")
        return

    # В DataFrame для удобства группировки
    df = pd.DataFrame(rows, columns=["Date", "VO2max"])
    df = df.sort_values("Date")

    daily = (
        df.groupby("Date", as_index=False)["VO2max"]
        .max()
        .sort_values("Date")
    )

    print("\nVO₂max (max/day) по месяцам:\n")

    current_month = None
    for _, row in daily.iterrows():
        d = row["Date"]
        v = row["VO2max"]
        key = (d.year, d.month)

        if key != current_month:
            if current_month is not None:
                print()
            print(f"{d.year}-{d.month:02d}")
            current_month = key

        print(f"{d}  {format_vo2(v)}")

    out_xlsx = "vo2max_history_daily.xlsx"
    wb = Workbook()

    ws_months = wb.active
    ws_months.title = "VO2max by month"

    monthly = {}
    for _, r in daily.iterrows():
        d = r["Date"]
        v = float(r["VO2max"])
        key = (d.year, d.month)
        monthly.setdefault(key, []).append((d, v))

    month_keys = sorted(monthly.keys())

    col_idx = 1

    for (year, month) in month_keys:
        dates_vo2 = monthly[(year, month)]
        month_label = f"{year}-{month:02d}"

        col_date = col_idx
        col_vo2 = col_idx + 1

        col_letter_date = get_column_letter(col_date)
        col_letter_vo2 = get_column_letter(col_vo2)

        cell_month = ws_months.cell(row=1, column=col_date, value=month_label)
        cell_month.font = Font(bold=True)

        ws_months.cell(row=2, column=col_date, value="Date").font = Font(bold=True)
        ws_months.cell(row=2, column=col_vo2, value="VO2max").font = Font(bold=True)

        row_idx = 3
        for d, v in dates_vo2:
            c_date = ws_months.cell(row=row_idx, column=col_date, value=d)
            c_date.number_format = "yyyy-mm-dd"

            c_vo2 = ws_months.cell(row=row_idx, column=col_vo2, value=v)
            c_vo2.number_format = "0.00"

            row_idx += 1

        ws_months.column_dimensions[col_letter_date].width = 12
        ws_months.column_dimensions[col_letter_vo2].width = 10

        col_idx += 3

    ws_chart = wb.create_sheet(title="VO2max Chart")

    ws_chart["A1"] = "Date"
    ws_chart["B1"] = "VO2max"
    ws_chart["A1"].font = Font(bold=True)
    ws_chart["B1"].font = Font(bold=True)

    row_idx = 2
    for _, r in daily.iterrows():
        d = r["Date"]
        v = float(r["VO2max"])

        c_date = ws_chart.cell(row=row_idx, column=1, value=d)
        c_date.number_format = "yyyy-mm-dd"

        c_vo2 = ws_chart.cell(row=row_idx, column=2, value=v)
        c_vo2.number_format = "0.00"

        row_idx += 1

    max_row = row_idx - 1

    ws_chart.column_dimensions["A"].width = 12
    ws_chart.column_dimensions["B"].width = 10

    chart = LineChart()
    chart.title = "VO2max over time"
    chart.y_axis.title = "ml/kg/min"
    chart.x_axis.title = "Date"
    chart.x_axis.number_format = "yyyy-mm-dd"

    data = Reference(ws_chart, min_col=2, min_row=1, max_row=max_row)
    cats = Reference(ws_chart, min_col=1, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws_chart.add_chart(chart, "D2")

    wb.save(out_xlsx)
    print(f"\nСохранено → {out_xlsx}")


if __name__ == "__main__":
    main()